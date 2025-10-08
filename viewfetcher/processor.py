"""Data ingestion helpers shared by the API layer."""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Dict, List, Optional, Tuple

from dateutil import parser as date_parser
from openpyxl import load_workbook

from .fetchers import extract_youtube_id, fetch_metrics, fetch_youtube_batch_stats

ALLOWED_PLATFORMS = {"youtube", "instagram", "tiktok"}


def _load_csv(file_bytes: bytes) -> List[Dict[str, object]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [{k: v for k, v in row.items()} for row in reader]


def _load_xlsx(file_bytes: bytes) -> List[Dict[str, object]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration as exc:  # pragma: no cover - guard against empty file
            raise ValueError("上传的文件为空") from exc

        normalized_headers = [str(h or "").strip() for h in headers]
        records: List[Dict[str, object]] = []
        for record in rows:
            values = {
                normalized_headers[i]: record[i] if i < len(record) else None
                for i in range(len(normalized_headers))
            }
            records.append(values)
        return records
    finally:
        workbook.close()


def _load_records(file_bytes: bytes, filename: str) -> List[Dict[str, object]]:
    if filename.lower().endswith(".csv"):
        rows = list(_load_csv(file_bytes))
    elif filename.lower().endswith(".xlsx"):
        rows = list(_load_xlsx(file_bytes))
    else:
        raise ValueError("仅支持 .csv 或 .xlsx 文件")

    if not rows:
        raise ValueError("上传的文件为空")

    normalized: List[Dict[str, object]] = []
    for index, raw in enumerate(rows, start=1):
        row = {str(k).strip().lower(): raw.get(k) for k in raw.keys()}
        url = str(row.get("url") or "").strip()
        if not url or not url.lower().startswith("http"):
            continue

        platform = str(row.get("platform") or "").strip().lower()
        if not platform:
            platform = _infer_platform(url)

        if platform not in ALLOWED_PLATFORMS:
            continue

        normalized_row = {
            "platform": platform,
            "url": url,
            "creator": _clean_text(row.get("creator")),
            "campaign_id": _clean_text(row.get("campaign_id")),
            "posted_at": row.get("posted_at"),
            "notes": _clean_text(row.get("notes")),
            "_row_number": index,
        }
        normalized.append(normalized_row)

    if not normalized:
        raise ValueError("未找到可识别的平台或链接")

    return normalized


def _infer_platform(url: str) -> str:
    url_l = url.lower()
    if "youtube.com" in url_l or "youtu.be" in url_l:
        return "youtube"
    if "instagram.com" in url_l:
        return "instagram"
    if "tiktok.com" in url_l:
        return "tiktok"
    return ""


def _clean_text(value: object) -> Optional[str]:
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return None if value is None else str(value)


def _parse_datetime(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    try:
        parsed = date_parser.parse(str(value))
    except (ValueError, TypeError):
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def process_file(file_bytes: bytes, filename: str, youtube_api_key: Optional[str]) -> Tuple[List[Dict], List[str]]:
    """Parse the uploaded file, fetch metrics and return results + error messages."""
    rows = _load_records(file_bytes, filename)

    results: List[Dict] = []
    errors: List[str] = []

    # YouTube batch (requires API key)
    yt_rows = [row for row in rows if row["platform"] == "youtube"]
    if yt_rows:
        if not youtube_api_key:
            errors.append("YouTube 数据未抓取：缺少 API Key")
        else:
            ids: List[str] = []
            indexed_rows: List[Tuple[str, Dict[str, object]]] = []
            for row in yt_rows:
                video_id = extract_youtube_id(row["url"])
                if video_id:
                    ids.append(video_id)
                    indexed_rows.append((video_id, row))

            stats_map: Dict[str, Dict] = {}
            chunk = 50
            for i in range(0, len(ids), chunk):
                batch = ids[i:i + chunk]
                try:
                    part = fetch_youtube_batch_stats(batch, youtube_api_key)
                    stats_map.update(part)
                except Exception as exc:  # pragma: no cover - network errors are runtime issues
                    errors.append(f"YouTube 抓取失败（{i + 1}-{i + len(batch)}）：{exc}")

            for video_id, row in indexed_rows:
                stats = stats_map.get(video_id, {})
                views = int(stats.get("views", 0))
                likes = int(stats.get("likes", 0))
                comments = int(stats.get("comments", 0))
                engagement_rate = round(((likes + comments) / views * 100.0), 2) if views > 0 else 0.0
                results.append({
                    "platform": "youtube",
                    "url": row["url"],
                    "creator": stats.get("creator") or row.get("creator"),
                    "campaign_id": row.get("campaign_id"),
                    "posted_at": _parse_datetime(stats.get("posted_at") or row.get("posted_at")),
                    "views": views,
                    "likes": likes,
                    "comments": comments,
                    "engagement_rate": engagement_rate,
                    "notes": row.get("notes"),
                })

    # Instagram & TikTok
    other_rows = [row for row in rows if row["platform"] in {"instagram", "tiktok"}]
    for idx, row in enumerate(other_rows, start=1):
        url = row["url"]
        platform = row["platform"]
        row_number = int(row.get("_row_number") or 0)
        try:
            stats = fetch_metrics(platform, url, youtube_api_key=None)
            views = int(stats.get("views", 0))
            likes = int(stats.get("likes", 0))
            comments = int(stats.get("comments", 0))
            engagement_rate = round(((likes + comments) / views * 100.0), 2) if views > 0 else 0.0
            results.append({
                "platform": platform,
                "url": url,
                "creator": stats.get("creator") or row.get("creator"),
                "campaign_id": row.get("campaign_id"),
                "posted_at": _parse_datetime(stats.get("posted_at") or row.get("posted_at")),
                "views": views,
                "likes": likes,
                "comments": comments,
                "engagement_rate": engagement_rate,
                "notes": row.get("notes"),
            })
        except Exception as exc:  # pragma: no cover - network errors at runtime
            hint = row_number if row_number else idx
            errors.append(f"{platform} 抓取失败（第 {hint} 行）：{exc}")

    return results, errors
